import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill


st.set_page_config(
    page_title="File Compare Tool",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# PWA Meta Tags
st.markdown("""
    <link rel="manifest" href="manifest.json">
    <meta name="theme-color" content="#667eea">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <meta name="apple-mobile-web-app-title" content="File Compare">
    <link rel="apple-touch-icon" href="logo.svg">
""", unsafe_allow_html=True)

# Custom CSS for better UI
st.markdown("""
    <style>
    /* Hide Streamlit menu and header */
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    .stDeployButton {display: none;}
    div[data-testid="stToolbar"] {display: none;}
    
    .header-container {
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 0.5rem;
        margin: 0.2rem 0 0.5rem 0;
        padding: 0;
    }
    .logo-container {
        display: flex;
        align-items: center;
        margin: 0;
    }
    .logo-container img {
        width: 32px;
        height: 32px;
        margin: 0;
    }
    .header-text {
        display: flex;
        flex-direction: column;
        align-items: flex-start;
        gap: 0;
    }
    .main-header {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-size: 1.4rem;
        font-weight: bold;
        margin: 0;
        padding: 0;
        line-height: 1.1;
    }
    .sub-header {
        color: #666;
        margin: 0;
        padding: 0;
        font-size: 0.75rem;
        line-height: 1.1;
    }
    .upload-section {
        background-color: #f8f9fa;
        padding: 1.5rem;
        border-radius: 10px;
        margin-bottom: 1rem;
    }
    .stButton>button {
        width: 100%;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        color: white;
        font-weight: bold;
        padding: 0.75rem;
        border-radius: 5px;
        border: none;
        font-size: 1.1rem;
    }
    .stButton>button:hover {
        background: linear-gradient(90deg, #764ba2 0%, #667eea 100%);
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
    }
    </style>
""", unsafe_allow_html=True)

# Logo and Header
import base64
try:
    with open("logo.svg", "rb") as f:
        logo_bytes = f.read()
        logo_base64 = base64.b64encode(logo_bytes).decode()
    st.markdown(f"""
        <div class="header-container">
            <div class="logo-container">
                <img src="data:image/svg+xml;base64,{logo_base64}" alt="Logo" style="width: 32px; height: 32px;">
            </div>
            <div class="header-text">
                <h1 class="main-header">File Compare Tool</h1>
                <p class="sub-header">Compare CSV and Excel files side-by-side</p>
            </div>
        </div>
    """, unsafe_allow_html=True)
except Exception as e:
    # Fallback to emoji if logo file not found
    st.markdown("""
        <div class="header-container">
            <div class="logo-container">
                <div style="font-size: 24px; margin: 0;">📊</div>
            </div>
            <div class="header-text">
                <h1 class="main-header">File Compare Tool</h1>
                <p class="sub-header">Compare CSV and Excel files side-by-side</p>
            </div>
        </div>
    """, unsafe_allow_html=True)

st.markdown("---")

col1, col2 = st.columns(2)

with col1:
    st.markdown("### 📄 File 1")
    uploaded1 = st.file_uploader(
        "Upload first file",
        type=["csv", "xlsx", "xls"],
        key="u1",
        help="Supported formats: CSV, Excel (.xlsx, .xls)"
    )

with col2:
    st.markdown("### 📄 File 2")
    uploaded2 = st.file_uploader(
        "Upload second file",
        type=["csv", "xlsx", "xls"],
        key="u2",
        help="Supported formats: CSV, Excel (.xlsx, .xls)"
    )

st.markdown("---")

col3, col4 = st.columns(2)

with col3:
    encoding = st.selectbox(
        "CSV Encoding",
        ["Auto", "utf-8", "cp1252", "latin1"],
        index=0,
        help="Select encoding for CSV files. Auto will try multiple encodings."
    )

with col4:
    show_only_diff = st.checkbox(
        "Show only differing rows",
        value=True,
        help="When enabled, only rows with differences will be displayed"
    )

st.markdown("---")

compare_button = st.button("🔍 Compare Files", use_container_width=True)

def detect_file_type(filename_or_path):
    """Detect if file is CSV or Excel based on extension."""
    if filename_or_path is None:
        return None
    path_str = str(filename_or_path).lower()
    if path_str.endswith(('.xlsx', '.xls')):
        return 'excel'
    elif path_str.endswith('.csv'):
        return 'csv'
    return None


def load_from_uploader(uploaded, enc_list):
    """Load file from uploader, supporting both CSV and Excel."""
    if uploaded is None:
        raise ValueError("Please upload a file")
    
    # Detect file type from uploaded file name
    file_type = detect_file_type(uploaded.name)
    
    if file_type == 'excel':
        try:
            uploaded.seek(0)
            df = pd.read_excel(uploaded, engine='openpyxl')
            return df, 'excel'
        except Exception as e:
            raise Exception(f"Failed to load Excel file: {e}")
    else:
        # CSV: Try encodings in order
        last_exc = None
        for enc in enc_list:
            try:
                uploaded.seek(0)
                df = pd.read_csv(uploaded, encoding=enc)
                return df, enc
            except Exception as e:
                last_exc = e
        raise last_exc or Exception("Failed to load CSV file with any encoding")


def normalize_value(val):
    """Normalize values for comparison: convert text numbers to numeric, keep strings as strings."""
    if pd.isna(val):
        return "<<NA>>"
    # Try to convert to number if it's a string representation of a number
    if isinstance(val, str):
        val_stripped = val.strip()
        if val_stripped == "":
            return ""
        # Try float first (handles both int and float strings)
        try:
            float_val = float(val_stripped)
            # If it's a whole number, return as int for consistency
            if float_val.is_integer():
                return int(float_val)
            return float_val
        except (ValueError, AttributeError):
            return val_stripped
    # If already numeric, normalize floats that are whole numbers to ints
    if isinstance(val, float) and val.is_integer():
        return int(val)
    return val


def normalize_dataframe_for_comparison(df):
    """Normalize all values in dataframe for comparison (text numbers -> numeric)."""
    df_norm = df.copy()
    for col in df_norm.columns:
        df_norm[col] = df_norm[col].apply(normalize_value)
    return df_norm


if compare_button:
    if uploaded1 is None or uploaded2 is None:
        st.error("⚠️ Please upload both files to compare")
        st.stop()
    
    encodings = [encoding] if encoding != "Auto" else ["utf-8", "cp1252", "latin1"]

    with st.spinner("Loading files..."):
        try:
            df1, file_info1 = load_from_uploader(uploaded1, encodings)
            if file_info1 == 'excel':
                st.success(f"✅ Loaded File 1: {uploaded1.name} (Excel)")
            else:
                st.success(f"✅ Loaded File 1: {uploaded1.name} (CSV, encoding: {file_info1})")
        except Exception as e:
            st.error(f"❌ Failed to load File 1: {e}")
            st.stop()

        try:
            df2, file_info2 = load_from_uploader(uploaded2, encodings)
            if file_info2 == 'excel':
                st.success(f"✅ Loaded File 2: {uploaded2.name} (Excel)")
            else:
                st.success(f"✅ Loaded File 2: {uploaded2.name} (CSV, encoding: {file_info2})")
        except Exception as e:
            st.error(f"❌ Failed to load File 2: {e}")
            st.stop()

    # Work on common columns to make highlight work predictably
    common_cols = df1.columns.intersection(df2.columns)
    if len(common_cols) == 0:
        st.warning("No common columns between the files — can't do a meaningful cell-by-cell comparison.")
        st.write("File 1 columns:", list(df1.columns))
        st.write("File 2 columns:", list(df2.columns))
        st.stop()

    df1c = df1[common_cols].copy()
    df2c = df2[common_cols].copy()

    # Align indices: reset to default range-based index for a row-wise comparison
    df1c = df1c.reset_index(drop=True)
    df2c = df2c.reset_index(drop=True)

    # Truncate to the smaller number of rows to allow pairwise comparison
    min_rows = min(len(df1c), len(df2c))
    if len(df1c) != len(df2c):
        st.info(f"Files have different number of rows ({len(df1c)} vs {len(df2c)}). Comparing first {min_rows} rows.")
    df1c = df1c.iloc[:min_rows]
    df2c = df2c.iloc[:min_rows]

    # Normalize dataframes for comparison (text numbers -> numeric)
    df1c_norm = normalize_dataframe_for_comparison(df1c)
    df2c_norm = normalize_dataframe_for_comparison(df2c)

    # Create mask of differences (treat NaN == NaN, text numbers == numeric)
    mask = ~(df1c_norm.eq(df2c_norm))
    differing_rows_mask = mask.any(axis=1)

    # Prepare filtered DataFrames if user wants only differing rows
    if show_only_diff:
        if not differing_rows_mask.any():
            st.info("No differing rows found in the compared range.")
        df1_display = df1c[differing_rows_mask].copy()
        df2_display = df2c[differing_rows_mask].copy()
    else:
        df1_display = df1c.copy()
        df2_display = df2c.copy()

    # --- Order-agnostic (multiset) comparison ---
    # Build hashable row keys from common columns (normalized: text numbers -> numeric, NaNs normalized)
    def rows_to_tuples(df: pd.DataFrame) -> pd.Series:
        # Use normalized dataframe for comparison
        df_norm = normalize_dataframe_for_comparison(df)
        # Convert to string representation for hashing (normalized values)
        return df_norm.astype(str).apply(lambda r: tuple(r.values.tolist()), axis=1)

    a_rows = rows_to_tuples(df1c)
    b_rows = rows_to_tuples(df2c)

    from collections import Counter
    ca = Counter(a_rows)
    cb = Counter(b_rows)

    # compute multiset symmetric difference count (number of row instances that don't match)
    all_keys = set(ca.keys()) | set(cb.keys())
    unordered_mismatch_count = sum(abs(ca.get(k, 0) - cb.get(k, 0)) for k in all_keys)

    st.markdown("---")
    st.markdown("## 📈 Comparison Results")
    
    col_metric1, col_metric2, col_metric3 = st.columns(3)
    with col_metric1:
        st.metric("Total Rows Compared", min_rows)
    with col_metric2:
        st.metric("Differing Rows", int(differing_rows_mask.sum()))
    with col_metric3:
        st.metric("Rows Mismatched (Order-agnostic)", unordered_mismatch_count)

    # Prepare DataFrames of unmatched rows (with repetitions equal to the difference in counts)
    def expand_counter_to_df(counter_source: Counter, counter_other: Counter, columns: pd.Index) -> pd.DataFrame:
        rows = []
        for key, cnt in counter_source.items():
            other_cnt = counter_other.get(key, 0)
            diff = cnt - other_cnt
            if diff > 0:
                for _ in range(diff):
                    rows.append(list(key))
        if rows:
            return pd.DataFrame(rows, columns=columns)
        return pd.DataFrame(columns=columns)

    unmatched_in_file1 = expand_counter_to_df(ca, cb, common_cols)
    unmatched_in_file2 = expand_counter_to_df(cb, ca, common_cols)

    if not unmatched_in_file1.empty or not unmatched_in_file2.empty:
        st.markdown("---")
        st.markdown("### 🔍 Unmatched Rows Summary")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f"**File 1** - {len(unmatched_in_file1)} unique rows")
            st.dataframe(unmatched_in_file1, use_container_width=True)
            if not unmatched_in_file1.empty:
                csv1 = unmatched_in_file1.to_csv(index=False).encode("utf-8")
                st.download_button("📥 Download File 1 Unmatched", data=csv1, file_name="unmatched_file1.csv", mime="text/csv", use_container_width=True)
        with c2:
            st.markdown(f"**File 2** - {len(unmatched_in_file2)} unique rows")
            st.dataframe(unmatched_in_file2, use_container_width=True)
            if not unmatched_in_file2.empty:
                csv2 = unmatched_in_file2.to_csv(index=False).encode("utf-8")
                st.download_button("📥 Download File 2 Unmatched", data=csv2, file_name="unmatched_file2.csv", mime="text/csv", use_container_width=True)

    # Show side-by-side with highlights
    st.markdown("---")
    st.markdown("### 🔄 Side-by-Side Comparison")
    st.markdown("*Differences are highlighted in red*")
    left, right = st.columns(2)

    def highlight_cols(col):
        return ["background-color: #ffcccc" if mask.at[idx, col.name] else "" for idx in col.index]

    with left:
        st.markdown(f"**📄 {uploaded1.name}**")
        try:
            st.dataframe(df1_display.style.apply(highlight_cols, axis=0), height=600, use_container_width=True)
        except Exception:
            st.dataframe(df1_display, height=600, use_container_width=True)

    with right:
        st.markdown(f"**📄 {uploaded2.name}**")
        try:
            st.dataframe(df2_display.style.apply(highlight_cols, axis=0), height=600, use_container_width=True)
        except Exception:
            st.dataframe(df2_display, height=600, use_container_width=True)

    # Show compact diff summary using pandas.compare
    st.markdown("---")
    st.markdown("### 📋 Compact Difference Report")
    try:
        # Compare on the displayed (possibly filtered) rows
        dfcomp = df1_display.compare(df2_display)
        if dfcomp.empty:
            st.success("✅ No differences found in compared rows/columns.")
        else:
            st.dataframe(dfcomp, use_container_width=True)
            # Allow download of the compact diff
            csv_bytes = dfcomp.to_csv(index=True).encode("utf-8")
            st.download_button("📥 Download Diff CSV", data=csv_bytes, file_name="diff.csv", mime="text/csv", use_container_width=True)

        # Also provide a combined CSV of differing rows with side-by-side values if the user chose only differing rows
        if show_only_diff and differing_rows_mask.any():
            combined = []
            for col in common_cols:
                combined.append(df1_display[col].rename(f"{col}_file1"))
                combined.append(df2_display[col].rename(f"{col}_file2"))
            combined_df = pd.concat(combined, axis=1)
            csv_bytes2 = combined_df.to_csv(index=True).encode("utf-8")
            st.download_button("📥 Download Side-by-Side CSV", data=csv_bytes2, file_name="differing_rows_side_by_side.csv", mime="text/csv", use_container_width=True)

            # Also offer an Excel download with differing cells highlighted in red
            try:
                excel_out = BytesIO()
                with pd.ExcelWriter(excel_out, engine="openpyxl") as writer:
                    combined_df.to_excel(writer, index=True, sheet_name="side_by_side")
                excel_out.seek(0)
                wb = openpyxl.load_workbook(excel_out)
                ws = wb["side_by_side"]
                red = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

                # Determine differing cells: columns are ordered as in combined_df
                # We'll compare file1/file2 pairs by column name suffix
                col_names = combined_df.columns.tolist()
                # find pairs like '<col>_file1' and '<col>_file2'
                pairs = []
                i = 0
                while i < len(col_names) - 1:
                    left = col_names[i]
                    right = col_names[i + 1]
                    if left.endswith("_file1") and right.endswith("_file2") and left[:-6] == right[:-6]:
                        pairs.append((i + 2, i + 3))  # +2 because openpyxl is 1-based and first column is index
                        i += 2
                    else:
                        i += 1

                # Apply red fill where values differ
                for (c1, c2) in pairs:
                    for r in range(2, ws.max_row + 1):
                        v1 = ws.cell(row=r, column=c1).value
                        v2 = ws.cell(row=r, column=c2).value
                        if (v1 is None and v2 is None) or (v1 == v2):
                            continue
                        ws.cell(row=r, column=c1).fill = red
                        ws.cell(row=r, column=c2).fill = red

                out2 = BytesIO()
                wb.save(out2)
                out2.seek(0)
                st.download_button("📥 Download Excel (Highlighted)", data=out2.getvalue(), file_name="differing_rows_side_by_side.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            except Exception as e:
                st.warning(f"Could not create Excel download: {e}")

        # Excel export: create a workbook with side-by-side sheet and compact diff
        if (show_only_diff and differing_rows_mask.any()) or (not show_only_diff):
            st.markdown("---")
            if st.button("📊 Export Full Report to Excel", use_container_width=True):
                out = BytesIO()
                with pd.ExcelWriter(out, engine="openpyxl") as writer:
                    # Write side-by-side if differing rows exist
                    if show_only_diff and differing_rows_mask.any():
                        combined_df.to_excel(writer, sheet_name="side_by_side", index=False)
                    else:
                        # If not filtering, write a side-by-side snapshot of compared rows
                        combined_all = []
                        for col in common_cols:
                            combined_all.append(df1c[col].rename(f"{col}_file1"))
                            combined_all.append(df2c[col].rename(f"{col}_file2"))
                        combined_all_df = pd.concat(combined_all, axis=1)
                        combined_all_df.to_excel(writer, sheet_name="side_by_side", index=False)

                    # Write compact diff (may be empty)
                    try:
                        comp_sheet = dfcomp if 'dfcomp' in locals() else df1c.compare(df2c)
                        comp_sheet.to_excel(writer, sheet_name="compact_diff")
                    except Exception:
                        # If compare fails, write a note
                        pd.DataFrame({"note": ["Could not compute compact diff"]}).to_excel(writer, sheet_name="compact_diff", index=False)

                # Apply highlighting to side_by_side sheet (reopen workbook in memory)
                out.seek(0)
                wb = openpyxl.load_workbook(out)
                if "side_by_side" in wb.sheetnames:
                    ws = wb["side_by_side"]
                    red = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

                    # Identify column pairs: columns are named like '<col>_file1', '<col>_file2'
                    max_col = ws.max_column
                    # Compare pairs (1-based column indices)
                    col = 1
                    while col < max_col:
                        for row in range(2, ws.max_row + 1):
                            cell1 = ws.cell(row=row, column=col)
                            cell2 = ws.cell(row=row, column=col + 1)
                            v1 = cell1.value
                            v2 = cell2.value
                            if (v1 is None and v2 is None) or (v1 == v2):
                                continue
                            # mark both cells
                            cell1.fill = red
                            cell2.fill = red
                        col += 2

                # Save workbook back to bytes
                out2 = BytesIO()
                wb.save(out2)
                out2.seek(0)
                st.download_button("📥 Download Excel Report", data=out2.getvalue(), file_name="diff_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    except Exception as e:
        st.error(f"❌ Could not compute compact diff: {e}")

    st.markdown("---")
    st.markdown("### 💡 Tips")
    st.info("""
    - **File Formats**: Supports CSV and Excel (.xlsx, .xls) files
    - **Text vs Numbers**: The tool automatically treats text numbers (e.g., "123") as equal to numeric values (123)
    - **Large Files**: For very large files, consider filtering to specific columns or rows before comparison
    - **Encoding**: If CSV files fail to load, try different encoding options
    """)
